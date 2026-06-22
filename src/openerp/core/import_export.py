from __future__ import annotations

import csv
import io
from collections.abc import Iterable
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import IO, Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.engine import Connection

from openerp.core.audit import log_operation
from openerp.core.context import RequestContext
from openerp.core.metadata import CatalogDef, FieldDef, FieldType, MetadataRegistry
from openerp.core.repository import Repository


def import_catalog_csv(
    connection: Connection,
    registry: MetadataRegistry,
    context: RequestContext,
    catalog_name: str,
    path: Path,
) -> list[int]:
    rows = read_csv_rows(path)
    result = import_catalog_rows(connection, registry, context, catalog_name, rows)
    return result["imported_ids"]


def read_csv_rows(source: Path | str | bytes | IO[bytes]) -> list[dict[str, str]]:
    if isinstance(source, (str, Path)):
        with open(source, encoding="utf-8", newline="") as handle:
            return list(csv.DictReader(handle))
    if isinstance(source, bytes):
        text = source.decode("utf-8-sig")
    elif hasattr(source, "read"):
        text = source.read().decode("utf-8-sig")
    else:
        text = str(source)
    return list(csv.DictReader(io.StringIO(text)))


def read_xlsx_rows(source: Path | str | bytes | IO[bytes]) -> list[dict[str, Any]]:
    if isinstance(source, (str, Path)):
        workbook = load_workbook(source, read_only=True, data_only=True)
    else:
        if isinstance(source, bytes):
            workbook = load_workbook(io.BytesIO(source), read_only=True, data_only=True)
        else:
            workbook = load_workbook(io.BytesIO(source.read()), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if headers is None:
        workbook.close()
        return []
    result = [dict(zip([str(h) for h in headers], row, strict=False)) for row in rows_iter]
    workbook.close()
    return result


def _coerce_catalog_value(field: FieldDef, raw: str | None) -> Any:
    if raw is None or raw == "":
        if field.required:
            raise ValueError(f"Missing required field: {field.name}")
        return None
    if field.type in (FieldType.INTEGER, FieldType.MONEY):
        return int(raw)
    if field.type == FieldType.BOOLEAN:
        return raw.lower() in ("1", "true", "yes", "on")
    return str(raw)


def import_catalog_rows(
    connection: Connection,
    registry: MetadataRegistry,
    context: RequestContext,
    catalog_name: str,
    rows: list[dict[str, str]],
    dry_run: bool = False,
) -> dict[str, Any]:
    catalog = _find_catalog(registry, catalog_name)
    repository = Repository(connection, registry, context)
    imported_ids: list[int] = []
    errors: list[dict[str, Any]] = []

    for index, row in enumerate(rows, start=2):
        try:
            name = (row.get("name") or "").strip()
            if not name:
                raise ValueError("Missing 'name' field")
            values: dict[str, Any] = {"name": name}
            for field in catalog.fields:
                value = _coerce_catalog_value(field, row.get(field.name))
                if value is not None:
                    values[field.name] = value
            if dry_run:
                continue
            item_id = repository.create_catalog_item(catalog_name, values)
            imported_ids.append(item_id)
        except Exception as exc:  # noqa: BLE001
            errors.append({"row": index, "error": str(exc)})

    if not dry_run and not errors:
        log_operation(
            connection,
            context,
            "import_catalog",
            object_type=f"catalog:{catalog_name}",
            details={"format": "rows", "count": len(rows)},
        )

    return {
        "imported": len(imported_ids),
        "errors": errors,
        "imported_ids": imported_ids,
    }


def import_initial_stock_rows(
    connection: Connection,
    registry: MetadataRegistry,
    context: RequestContext,
    rows: list[dict[str, str]],
    dry_run: bool = False,
) -> dict[str, Any]:
    """Create and post an inventory_adjustment document from CSV/XLSX rows.

    Expected columns: product_sku, warehouse_name, quantity.
    """
    product_map = _catalog_lookup_by_key(
        connection, "product", "sku", context.organization_id
    )
    warehouse_map = _catalog_lookup_by_key(
        connection, "warehouse", "name", context.organization_id
    )
    errors: list[dict[str, Any]] = []
    lines_by_warehouse: dict[int, list[dict[str, Any]]] = defaultdict(list)

    for index, row in enumerate(rows, start=2):
        sku = (row.get("product_sku") or "").strip()
        wh = (row.get("warehouse_name") or "").strip()
        qty_raw = (row.get("quantity") or "").strip()
        if not sku or not wh or not qty_raw:
            errors.append(
                {"row": index, "error": "Missing product_sku, warehouse_name or quantity"}
            )
            continue
        product_id = product_map.get(sku)
        if product_id is None:
            errors.append({"row": index, "error": f"Unknown product SKU: {sku}"})
            continue
        warehouse_id = warehouse_map.get(wh)
        if warehouse_id is None:
            errors.append({"row": index, "error": f"Unknown warehouse: {wh}"})
            continue
        try:
            quantity = str(qty_raw)
        except (TypeError, ValueError):
            errors.append({"row": index, "error": f"Invalid quantity: {qty_raw}"})
            continue
        lines_by_warehouse[warehouse_id].append(
            {"product_id": product_id, "quantity_delta": quantity}
        )

    if errors or dry_run or not lines_by_warehouse:
        return {"imported": 0, "errors": errors, "document_id": None, "document_ids": []}

    repository = Repository(connection, registry, context)
    from openerp.core.posting import DocumentPostingService

    poster = DocumentPostingService(connection, registry, context)
    document_ids: list[int] = []
    imported_count = 0
    for warehouse_id, warehouse_lines in lines_by_warehouse.items():
        document_id = repository.create_document(
            "inventory_adjustment",
            {"date": date.today(), "warehouse_id": warehouse_id},
            {"lines": warehouse_lines},
        )
        poster.post("inventory_adjustment", document_id)
        document_ids.append(document_id)
        imported_count += len(warehouse_lines)

    log_operation(
        connection,
        context,
        "import_initial_stock",
        object_type="document:inventory_adjustment",
        details={"document_ids": document_ids, "line_count": imported_count},
    )
    return {
        "imported": imported_count,
        "errors": errors,
        "document_id": document_ids[0] if len(document_ids) == 1 else None,
        "document_ids": document_ids,
    }


def catalog_template_rows(registry: MetadataRegistry, catalog_name: str) -> list[dict[str, str]]:
    catalog = _find_catalog(registry, catalog_name)
    headers = ["name"] + [field.name for field in catalog.fields]
    template_row = {header: "" for header in headers}
    for field in catalog.fields:
        if field.default is not None:
            template_row[field.name] = str(field.default)
    return [template_row]


def _find_catalog(registry: MetadataRegistry, catalog_name: str) -> CatalogDef:
    for catalog in registry.catalogs():
        if catalog.name == catalog_name:
            return catalog
    raise ValueError(f"Unknown catalog: {catalog_name}")


def _catalog_lookup_by_key(
    connection: Connection,
    catalog_name: str,
    key_field: str,
    organization_id: int,
) -> dict[str, int]:
    from openerp.core.naming import catalog_table

    table = connection.engine._openerp_metadata.tables[catalog_table(catalog_name)]
    rows = connection.execute(
        select(table.c.id, getattr(table.c, key_field)).where(
            table.c.organization_id == organization_id,
            table.c.deletion_mark.is_(False),
        )
    )
    return {row._mapping[key_field]: row._mapping["id"] for row in rows if row._mapping[key_field]}


def _write_csv(rows: list[dict], handle: IO[str]) -> None:
    if not rows:
        return
    writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)


def export_rows_csv(rows: Iterable[dict], target: Path | str | IO[bytes]) -> None:
    rows = list(rows)
    if isinstance(target, (str, Path)):
        with open(target, "w", encoding="utf-8", newline="") as handle:
            _write_csv(rows, handle)
    else:
        wrapper = io.TextIOWrapper(target, encoding="utf-8", newline="", write_through=True)
        try:
            _write_csv(rows, wrapper)
        finally:
            wrapper.detach()


def export_rows_xlsx(rows: Iterable[dict], target: Path | str | IO[bytes]) -> None:
    rows = list(rows)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"
    if rows:
        headers = list(rows[0].keys())
        worksheet.append(headers)
        for row in rows:
            worksheet.append([row.get(header) for header in headers])
    workbook.save(target)
